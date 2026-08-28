# backend/main.py
from datetime import datetime
from typing import Any, Dict, List, Optional

from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware

from io import BytesIO
from fastapi.responses import StreamingResponse
import openpyxl
import json
from collections import Counter


from sqlalchemy import (
    Column, Integer, String, JSON, ForeignKey, create_engine, select, delete
)
from sqlalchemy.orm import declarative_base, sessionmaker, Session

# ---------- DB setup ----------
DB_URL = "sqlite:///data.db"
engine = create_engine(DB_URL, echo=False, future=True)
SessionLocal = sessionmaker(bind=engine, autoflush=False, autocommit=False, future=True)
Base = declarative_base()

# ---------- Tables ----------
class Response(Base):
    __tablename__ = "responses"
    id = Column(Integer, primary_key=True, index=True)
    response_id = Column(Integer, nullable=True)
    ts = Column(String, nullable=False)
    payload = Column(JSON, nullable=False)

class Question(Base):
    __tablename__ = "questions"
    id = Column(Integer, primary_key=True, index=True)
    response_id = Column(Integer, nullable=True)
    text = Column(String, nullable=False)
    qtype = Column(String, nullable=False, default="single")   # single | multi | text
    qorder = Column(Integer, nullable=False, default=0)

class Option(Base):
    __tablename__ = "options"
    id = Column(Integer, primary_key=True, index=True)
    response_id = Column(Integer, nullable=True)
    question_id = Column(Integer, ForeignKey("questions.id"), nullable=False)
    code = Column(String, nullable=False)
    label = Column(String, nullable=False)
    oorder = Column(Integer, nullable=False, default=0)


# ---------- Tables ----------
class TrackingData(Base):
    __tablename__ = "tracking_data"

    id = Column(Integer, primary_key=True, index=True)
    response_id = Column(Integer, nullable=True)
    recording = Column(String, nullable=True)
    participant = Column(String, nullable=True)
    toi = Column(String, nullable=True)
    interval = Column(String, nullable=True)
    aoi = Column(String, nullable=True)

    duration_of_interval = Column(String, nullable=True)
    total_duration_of_fixations = Column(String, nullable=True)
    average_duration_of_fixations = Column(String, nullable=True)
    minimum_duration_of_fixations = Column(String, nullable=True)
    maximum_duration_of_fixations = Column(String, nullable=True)

    number_of_fixations = Column(Integer, nullable=True)
    number_of_mouse_clicks = Column(Integer, nullable=True)
    time_to_first_mouse_click = Column(String, nullable=True)
    number_of_touch_events = Column(Integer, nullable=True)

    time_to_first_touch = Column(String, nullable=True)
    device = Column(String, nullable=True)

Base.metadata.create_all(engine)    
# ---------- App ----------
app = FastAPI(title="Rail Survey API")

tracking_buffer = {
    "mouse_clicks": 0,
    "touch_events": 0,
    "first_mouse_click": None,
    "first_touch": None,
    "device": "unknown"
}

app.add_middleware(
    CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"]
)

def db() -> Session:
    s = SessionLocal()
    try:
        yield s
    finally:
        s.close()

# ---------- Health ----------
@app.get("/health")
def health(): return {"ok": True}

# ---------- Responses ----------
@app.post("/submit")
def submit(payload: Dict[str, Any]):
    if not isinstance(payload, dict):
        raise HTTPException(400, "payload must be a JSON object")

    with SessionLocal() as s:
        r = Response(
            ts=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            payload=payload
        )

        s.add(r)
        s.commit()
        s.refresh(r)

    return {
        "ok": True,
        "response_id": r.id
    }

@app.get("/responses")
def responses() -> List[Dict[str, Any]]:
    with SessionLocal() as s:
        rows = s.execute(select(Response).order_by(Response.id)).scalars().all()
        return [{"id":r.id, "ts":r.ts, "payload":r.payload} for r in rows]
    
# ---------- Tracking Data ----------
@app.post("/tracking")
def save_tracking(payload: Dict[str, Any]):
    with SessionLocal() as s:
        tracking = TrackingData(
            response_id=payload.get("response_id"),
            recording=payload.get("Recording"),
            participant=payload.get("Participant"),
            toi=payload.get("TOI"),
            interval=payload.get("Interval"),
            aoi=payload.get("AOI"),
            duration_of_interval=payload.get("Duration_of_interval"),
            total_duration_of_fixations=payload.get("Total_duration_of_fixations"),
            average_duration_of_fixations=payload.get("Average_duration_of_fixations"),
            minimum_duration_of_fixations=payload.get("Minimum_duration_of_fixations"),
            maximum_duration_of_fixations=payload.get("Maximum_duration_of_fixations"),
            number_of_fixations=payload.get("Number_of_fixations"),
            number_of_mouse_clicks=payload.get("Number_of_mouse_clicks"),
            time_to_first_mouse_click=payload.get("Time_to_first_mouse_click"),
            number_of_touch_events=payload.get("Number_of_touch_events"),
            time_to_first_touch=payload.get("Time_to_first_touch"),
            device=payload.get("Device_type"),
        )

        s.add(tracking)
        s.commit()

    return {"ok": True}
# ---------- Questions CRUD ----------
# ساخت سؤال
@app.post("/question")
def create_question(q: Dict[str, Any]):
    """
    payload نمونه:
    {
      "text": "Which station are you at today?",
      "qtype": "single",
      "qorder": 0,
      "options": [
        {"code":"do","label":"Dortmund Hbf"},
        {"code":"es","label":"Essen Hbf"}
      ]
    }
    """
    text = q.get("text"); qtype = q.get("qtype","single"); qorder = q.get("qorder",0)
    if not text: raise HTTPException(400,"text is required")

    with SessionLocal() as s:
        qrow = Question(text=text, qtype=qtype, qorder=qorder)
        s.add(qrow); s.commit(); s.refresh(qrow)

        opts = q.get("options") or []
        for i, o in enumerate(opts):
            s.add(Option(question_id=qrow.id,
                         code=o.get("code", f"opt{i}"),
                         label=o.get("label",""),
                         oorder=i))
        s.commit()

        return {"id": qrow.id}

# گرفتن همه سؤال‌ها (برای فرانت)
@app.get("/questions")
def get_questions():
    with SessionLocal() as s:
        qrows = s.execute(select(Question).order_by(Question.qorder, Question.id)).scalars().all()
        out = []
        for q in qrows:
            opts = s.execute(
                select(Option).where(Option.question_id==q.id).order_by(Option.oorder, Option.id)
            ).scalars().all()
            out.append({
                "id": q.id, "text": q.text, "type": q.qtype, "order": q.qorder,
                "options": [{"id":o.id,"code":o.code,"label":o.label,"order":o.oorder} for o in opts]
            })
        return {"questions": out}

#.

# حذف سؤال
@app.delete("/question/{qid}")
def delete_question(qid: int):
    with SessionLocal() as s:
        s.execute(delete(Option).where(Option.question_id==qid))
        row = s.get(Question, qid)
        if row: s.delete(row)
        s.commit()
    return {"ok": True}

# ---------- Update Question ----------
@app.put("/question/{qid}")
def update_question(qid: int, payload: Dict[str, Any]):

    with SessionLocal() as s:

        qrow = s.get(Question, qid)

        if not qrow:
            raise HTTPException(404, "Question not found")


        # update question fields
        qrow.text = payload.get("text", qrow.text)
        qrow.qtype = payload.get("qtype", qrow.qtype)
        qrow.qorder = payload.get("qorder", qrow.qorder)


        # delete old options
        s.execute(
            delete(Option).where(
                Option.question_id == qid
            )
        )


        # add new options
        options = payload.get("options") or []

        for i, o in enumerate(options):

            s.add(
                Option(
                    question_id=qid,
                    code=o.get("code", f"option_{i+1}"),
                    label=o.get("label", ""),
                    oorder=i
                )
            )


        s.commit()

        return {
            "ok": True,
            "id": qid
        }
# ---------- Excel Export ----------
from io import BytesIO
from fastapi.responses import StreamingResponse
import openpyxl, json
from collections import Counter
from sqlalchemy import select

@app.get("/export.xlsx")
def export_excel():
    # 1) داده‌ها از DB
    with SessionLocal() as s:
        qrows = s.execute(select(Question).order_by(Question.qorder, Question.id)).scalars().all()
        orows = s.execute(select(Option).order_by(Option.oorder, Option.id)).scalars().all()
        rrows = s.execute(select(Response).order_by(Response.id)).scalars().all()

    # 2) map گزینه‌ها: {qid: {code: label}}
    optmap = {}
    for o in orows:
        optmap.setdefault(o.question_id, {})[o.code] = o.label

    # 3) ساخت Workbook
    wb = openpyxl.Workbook()

    # Sheet خام
    ws_raw = wb.active
    ws_raw.title = "raw_responses"
    ws_raw.append(["id", "ts", "payload"])
    for r in rrows:
        ws_raw.append([r.id, r.ts, json.dumps(r.payload, ensure_ascii=False)])

    # Sheet خلاصه کل
    ws_sum = wb.create_sheet("summary_counts")
    ws_sum.append(["Question", "Option", "Count"])

    # شیت برای هر سؤال
    for q in qrows:
        qid, qtext, qtype = q.id, q.text, q.qtype

        if qtype in ("single", "multi"):
            counts = Counter()
            for r in rrows:
                ans = (r.payload or {}).get("answers", {})
                v = ans.get(str(qid)) if isinstance(ans, dict) else None
                if v is None:
                    continue
                if qtype == "multi" and isinstance(v, list):
                    for code in v:
                        counts[code] += 1
                else:
                    counts[v] += 1

            ws = wb.create_sheet(f"Q{qid}_counts")
            ws.append(["Question", "Option", "Count"])
            for code, n in counts.items():
                label = optmap.get(qid, {}).get(code, str(code))
                ws.append([qtext, label, n])
                ws_sum.append([qtext, label, n])

        else:  # text
            ws = wb.create_sheet(f"Q{qid}_texts")
            ws.append(["response_id", "ts", "Answer"])
            for r in rrows:
                ans = (r.payload or {}).get("answers", {})
                v = ans.get(str(qid)) if isinstance(ans, dict) else None
                if v is not None:
                    ws.append([r.id, r.ts, str(v)])

    # 4) ارسال فایل به‌صورت استریم
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    headers = {"Content-Disposition": 'attachment; filename="survey_export.xlsx"'}
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )

#-----------------------------------------------------
@app.post("/tracking_event")
def tracking_event(payload: Dict[str, Any]):
    print(payload)
    event = payload.get("event")

    if event == "click":
        tracking_buffer["mouse_clicks"] += 1

        if tracking_buffer["first_mouse_click"] is None:
            tracking_buffer["first_mouse_click"] = payload.get("first_mouse_click")


    elif event == "touch":
        tracking_buffer["touch_events"] += 1

        if tracking_buffer["first_touch"] is None:
            tracking_buffer["first_touch"] = payload.get("first_touch")


    tracking_buffer["device"] = payload.get(
        "device",
        "unknown"
    )


    return {
        "ok": True
    }


@app.get("/tracking_latest")
def tracking_latest():
    with SessionLocal() as s:
        row = s.execute(
            select(TrackingData)
            .order_by(TrackingData.id.desc())
        ).scalars().first()

        if not row:
            return {"message": "No tracking data"}

        return {
            "id": row.id,
            "clicks": row.number_of_mouse_clicks,
            "first_click": row.time_to_first_mouse_click,
            "interval": row.interval,
            "recording": row.recording
        }
# ---------- Complete Survey Export ----------
@app.get("/export_complete.xlsx")
def export_complete_excel():

    with SessionLocal() as s:

        questions = s.execute(
            select(Question).order_by(Question.qorder, Question.id)
        ).scalars().all()

        options = s.execute(
            select(Option)
        ).scalars().all()

        responses = s.execute(
            select(Response).order_by(Response.id)
        ).scalars().all()
       

    # تبدیل option code به label
    option_map = {}

    for o in options:
        option_map.setdefault(o.question_id, {})
        option_map[o.question_id][o.code] = o.label


    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "complete_survey"


    # Header
    headers = [
        "Participant",
        "Timestamp"
    ]


    for q in questions:
        headers.append(
            f"Q{q.id}: {q.text}"
        )


    # Tracking columns
    headers.extend([
    "Recording",
    "TOI",
    "Interval",
    "AOI",
    "Duration_of_interval",
    "Number_of_mouse_clicks",
    "Time_to_first_mouse_click",
    "Number_of_touch_events",
    "Time_to_first_touch",
    "Device_type"
    ])


    ws.append(headers)


    # Responses
    for r in responses:

        row = [
            r.id,
            r.ts
        ]
        
        tracking = s.execute(
            select(TrackingData)
            .where(TrackingData.response_id == r.id)
        ).scalars().first()

        answers = (r.payload or {}).get(
            "answers",
            {}
        )


        for q in questions:

            value = answers.get(
                str(q.id),
                ""
            )


            # تبدیل code به label
            if isinstance(value, list):

                labels = [
                    option_map.get(q.id, {}).get(
                        x,
                        x
                    )
                    for x in value
                ]

                value = ", ".join(labels)


            else:

                value = option_map.get(
                    q.id,
                    {}
                ).get(
                    value,
                    value
                )


            row.append(value)


        # Tracking آخرین رکورد
        if tracking:

           row.extend([
                tracking.recording,
                tracking.toi,
                tracking.interval,
                tracking.aoi,
                tracking.duration_of_interval,
                tracking.number_of_mouse_clicks,
                tracking.time_to_first_mouse_click,
                tracking.number_of_touch_events,
                tracking.time_to_first_touch,
                tracking.device
            ])

        else:

            row.extend([
                "",
                "",
                "",
                "",
                "",
                "",
                ""
            ])


        ws.append(row)



    # Auto width
    for col in ws.columns:

        length = max(
            len(str(cell.value))
            if cell.value else 0
            for cell in col
        )

        ws.column_dimensions[
            col[0].column_letter
        ].width = min(length + 3, 50)



    buf = BytesIO()

    wb.save(buf)

    buf.seek(0)


    return StreamingResponse(
        buf,
        media_type=
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition":
            'attachment; filename="CNY_Survey_Complete_Data.xlsx"'
        }
    )
#..................................................................................
@app.delete("/clear_data")
def clear_data():

    with SessionLocal() as s:

        s.execute(delete(Response))
        s.execute(delete(TrackingData))

        s.commit()

    return {
        "ok": True,
        "message": "All responses and tracking data deleted"
    }
#..............................................................................
@app.get("/tracking_buffer")
def get_tracking_buffer():
    return tracking_buffer